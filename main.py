from fastapi import FastAPI
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional
from openpyxl import load_workbook
from datetime import datetime
import io, os, httpx, subprocess, tempfile

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

MODELO_URL = "https://lbhjtlfryysjdbavlkgm.supabase.co/storage/v1/object/public/modelos-excel/modelo_in_situ.xlsx"

class DadosEnsaio(BaseModel):
    amostra: Optional[str] = ""
    camada: Optional[str] = ""
    tipoMaterial: Optional[str] = ""
    densLab: Optional[float] = 0
    umidadeOtima: Optional[float] = 0
    metodoUmidade: Optional[str] = "direto"
    frigideiraUmido: Optional[float] = 0
    frigideiraSeco: Optional[float] = 0
    tara: Optional[float] = 0
    frascoAntes: Optional[float] = 0
    frascoDepois: Optional[float] = 0
    funil: Optional[float] = 0
    densAreia: Optional[float] = 0
    amostraUmida: Optional[float] = 0
    ident: Optional[str] = ""
    cliente: Optional[str] = ""
    construtora: Optional[str] = ""
    data: Optional[str] = ""
    gcMin: Optional[float] = None
    desvioMax: Optional[float] = None
    observacao: Optional[str] = ""

class PayloadExcel(BaseModel):
    lista_ensaios: List[DadosEnsaio]
    relatorio_obs: Optional[str] = ""
    formato: Optional[str] = "excel"  # "excel" ou "pdf"

def preencher_workbook(wb, lista, relatorio_obs):
    ws = wb.worksheets[0]
    COLUNAS = ['C','D','E','F','G','H','I','J','K']

    def set_cell(addr, value):
        for merge in ws.merged_cells.ranges:
            if addr in merge:
                ws.cell(row=merge.min_row, column=merge.min_col).value = value
                return
        ws[addr].value = value

    d0 = lista[0]

    # Cabeçalho
    ws['B6'].value  = f"OBRA: {d0.ident or ''}"
    ws['C7'].value  = d0.cliente or ''
    ws['B8'].value  = f"CONSTRUTORA: {d0.construtora or ''}"
    ws['B9'].value  = f"TRECHO: {d0.amostra or ''}"
    ws['B10'].value = f"CAMADA: {d0.camada or ''}"
    ws['B11'].value = f"LOCAL: {d0.ident or ''}"

    if d0.data:
        try:
            ws['J4'].value = datetime.strptime(d0.data, '%Y-%m-%d')
        except:
            ws['J4'].value = d0.data

    # OBS — prioriza obs do técnico, depois do Claude
    obs_final = ''
    if d0.observacao:
        obs_final = d0.observacao
    if relatorio_obs:
        obs_final = (obs_final + '\n' + relatorio_obs).strip() if obs_final else relatorio_obs
    if obs_final:
        ws['C51'].value = obs_final

    for idx, d in enumerate(lista[:9]):
        c = COLUNAS[idx]
        num = idx + 1

        ws[f'{c}14'].value = num
        if d.data:
            try:
                ws[f'{c}15'].value = datetime.strptime(d.data, '%Y-%m-%d')
            except:
                ws[f'{c}15'].value = d.data

        ws[f'{c}16'].value = d.amostra or ''
        ws[f'{c}17'].value = 1
        ws[f'{c}18'].value = d.camada or ''
        ws[f'{c}19'].value = d.tipoMaterial or ''
        ws[f'{c}20'].value = 0.2

        dens_cell = ws[f'{c}22']
        dens_cell.value = int(round(d.densLab or 0))
        dens_cell.number_format = '0'
        ws[f'{c}23'].value = d.umidadeOtima or 0

        if d.metodoUmidade == 'frigideira':
            ws[f'{c}25'].value = d.frigideiraUmido or 0
            ws[f'{c}26'].value = d.frigideiraSeco or 0
            ws[f'{c}27'].value = d.tara or 0

        ws[f'{c}28'].value = f'={c}25-{c}27'
        ws[f'{c}29'].value = f'={c}26-{c}27'
        ws[f'{c}30'].value = f'={c}28-{c}29'
        ws[f'{c}31'].value = f'=IF({c}29=0,0,{c}30/{c}29*100)'
        ws[f'{c}33'].value = d.frascoAntes or 0
        ws[f'{c}34'].value = d.frascoDepois or 0
        ws[f'{c}35'].value = f'={c}33-{c}34'
        ws[f'{c}36'].value = d.funil or 0
        ws[f'{c}37'].value = f'={c}35-{c}36'
        ws[f'{c}38'].value = d.densAreia or 0
        ws[f'{c}39'].value = f'=IF({c}38=0,0,{c}37/{c}38)'
        ws[f'{c}40'].value = d.amostraUmida or 0
        ws[f'{c}41'].value = 0
        ws[f'{c}42'].value = 0
        ws[f'{c}43'].value = f'=IF({c}41=0,0,{c}41/{c}42)'
        ws[f'{c}44'].value = f'=IF({c}41=0,{c}40,{c}40-{c}41)'
        ws[f'{c}45'].value = f'=IF({c}41=0,{c}39,{c}39-{c}43)'
        ws[f'{c}46'].value = f'=IF({c}45=0,0,{c}44/{c}45)'
        ws[f'{c}47'].value = f'=IF({c}31=0,{c}46,{c}46/(1+({c}31/100)))'
        ws[f'{c}48'].value = f'=IF({c}22=0,0,({c}47/{c}22)*100)'
        ws[f'{c}49'].value = f'={c}31-{c}23'

    return ws, wb

@app.get("/")
def health():
    return {"status": "ok", "servico": "LabCQ Excel/PDF API v3"}

@app.post("/gerar-excel")
async def gerar_excel(payload: PayloadExcel):
    async with httpx.AsyncClient() as client:
        resp = await client.get(MODELO_URL)
        if resp.status_code != 200:
            return JSONResponse({"erro": f"Modelo não encontrado: {resp.status_code}"}, status_code=500)

    wb = load_workbook(io.BytesIO(resp.content))
    ws, wb = preencher_workbook(wb, payload.lista_ensaios, payload.relatorio_obs)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    d0 = payload.lista_ensaios[0]
    nome = f"InSitu_{(d0.amostra or d0.ident or 'ensaio').replace(' ','_')}"

    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{nome}.xlsx"'}
    )

@app.post("/gerar-pdf")
async def gerar_pdf(payload: PayloadExcel):
    async with httpx.AsyncClient() as client:
        resp = await client.get(MODELO_URL)
        if resp.status_code != 200:
            return JSONResponse({"erro": f"Modelo não encontrado: {resp.status_code}"}, status_code=500)

    wb = load_workbook(io.BytesIO(resp.content))
    ws, wb = preencher_workbook(wb, payload.lista_ensaios, payload.relatorio_obs)

    d0 = payload.lista_ensaios[0]
    nome = f"InSitu_{(d0.amostra or d0.ident or 'ensaio').replace(' ','_')}"

    # Salva Excel temporário
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False, dir='/tmp') as tmp:
        wb.save(tmp.name)
        xlsx_path = tmp.name

    # Converte para PDF com LibreOffice
    result = subprocess.run([
        'libreoffice', '--headless', '--convert-to', 'pdf',
        '--outdir', '/tmp', xlsx_path
    ], capture_output=True, text=True, timeout=60)

    pdf_path = xlsx_path.replace('.xlsx', '.pdf')

    if not os.path.exists(pdf_path):
        return JSONResponse({"erro": f"Falha na conversão: {result.stderr}"}, status_code=500)

    with open(pdf_path, 'rb') as f:
        pdf_bytes = f.read()

    # Limpa arquivos temporários
    os.unlink(xlsx_path)
    os.unlink(pdf_path)

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{nome}.pdf"'}
    )
